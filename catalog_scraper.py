#!/usr/bin/env python3
"""
Catalog Scraper — Finds Steel City Vacuum products NOT in the schematics.

Strategy:
  1. Load existing part numbers from steel_city_parts.xlsx (12,576 known)
  2. Login via undetected-chromedriver (Cloudflare bypass)
  3. Navigate to the Store/Catalog section (/a/s/)
  4. Discover all product categories from the page
  5. Scrape each category page (with pagination) for product listings
  6. Supplement with search API using broad product terms
  7. For each NEW product, call product_info API for full details
  8. Filter out: special order items, NLA without alternatives
  9. Export new products to Excel

Usage:
  python3 catalog_scraper.py                    # Full pipeline
  python3 catalog_scraper.py --step discover    # Only discover categories + products
  python3 catalog_scraper.py --step enrich      # Only enrich already-discovered products
  python3 catalog_scraper.py --step export      # Only export to Excel
"""

import argparse
import json
import logging
import os
import random
import re
import time
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException,
    WebDriverException
)
import undetected_chromedriver as uc

# ── Config ──────────────────────────────────────────────────────────────────

CONFIG = {
    "base_url": "https://www.steelcityvac.com",
    "store_url": "https://www.steelcityvac.com/a/s/",
    "account": "REDACTED_ACCT",
    "user_id": "REDACTED_USER",
    "password": "REDACTED_PASS",
}

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
SPREADSHEET = OUTPUT_DIR / "steel_city_parts.xlsx"
PROGRESS_FILE = BASE_DIR / "catalog_progress.json"
DEBUG_DIR = BASE_DIR / "debug"
DEBUG_DIR.mkdir(exist_ok=True)

IMAGE_BASE_URL = "https://www.steelcityvac.com/uploads/applications/shopping_cart/"

# Search terms to find non-schematic products (bags, cleaners, whole units, etc.)
SEARCH_TERMS = [
    # Product types
    "bag", "bags", "filter", "belt", "brush", "hose", "cord", "switch",
    "motor", "fan", "wheel", "roller", "nozzle", "wand", "attachment",
    "tool", "crevice", "upholstery", "dusting", "extension",
    # Consumables / cleaning products
    "cleaner", "cleaning", "deodorizer", "freshener", "scent", "tablet",
    "solution", "polish", "spray", "wipe",
    # Whole units / types
    "vacuum", "upright", "canister", "backpack", "stick", "handheld",
    "portable", "commercial", "central",
    # HEPA / filtration
    "hepa", "carbon", "foam", "exhaust", "pre-filter", "post-filter",
    # Common brands (to catch brand-specific accessories)
    "dyson", "hoover", "eureka", "bissell", "shark", "kirby", "miele",
    "sebo", "riccar", "oreck", "electrolux", "panasonic", "kenmore",
    "royal", "sanitaire", "proteam", "windsor", "karcher", "nilfisk",
    "clarke", "advance", "tornado", "pacific", "powr-flite", "nss",
    "lindhaus", "cirrus", "simplicity", "fuller", "compact", "tristar",
    "aerus", "nutone", "beam", "md", "cyclovac", "drainvac", "duo-vac",
    "hayden", "vacuflo", "honeywell", "broan",
    # Part number prefixes (catch numeric-only parts)
    "0", "1", "2", "3", "4", "5", "6", "7", "8", "9",
    # Letters for alphanumeric parts
    "a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m",
    "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z",
]

CHECKPOINT_INTERVAL = 25

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("catalog_scraper")


# ── Helpers ─────────────────────────────────────────────────────────────────

def random_delay(lo=1.0, hi=2.5):
    time.sleep(random.uniform(lo, hi))


def human_type(element, text, delay_range=(0.05, 0.12)):
    element.click()
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(*delay_range))


def load_existing_part_numbers():
    """Load all part numbers from the schematics spreadsheet."""
    log.info(f"Loading existing parts from {SPREADSHEET}...")
    wb = openpyxl.load_workbook(str(SPREADSHEET), read_only=True)
    ws = wb.active
    part_numbers = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row[3] or "").strip()
        if pn:
            part_numbers.add(pn)
    wb.close()
    log.info(f"  Loaded {len(part_numbers)} existing part numbers")
    return part_numbers


def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {
        "categories_discovered": [],
        "categories_scraped": [],
        "search_terms_done": [],
        "discovered_products": {},  # part_number -> {basic info}
        "enriched_products": {},    # part_number -> {full info from API}
        "skipped_existing": 0,
        "skipped_special_order": 0,
        "skipped_nla": 0,
    }


def save_progress(progress):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def screenshot(driver, name):
    path = str(DEBUG_DIR / f"catalog_{name}.png")
    try:
        driver.save_screenshot(path)
    except Exception:
        pass


def save_html(driver, name):
    path = str(DEBUG_DIR / f"catalog_{name}.html")
    try:
        with open(path, "w") as f:
            f.write(driver.page_source)
    except Exception:
        pass


# ── Browser ─────────────────────────────────────────────────────────────────

def create_driver():
    options = uc.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument(f"--user-data-dir={BASE_DIR / 'browser_data'}")
    return uc.Chrome(options=options, headless=False, version_main=145)


def wait_for_cloudflare(driver, max_wait=120):
    if "just a moment" not in driver.title.lower():
        return
    log.info("Cloudflare challenge detected — waiting...")
    for i in range(max_wait // 2):
        time.sleep(2)
        if "just a moment" not in driver.title.lower():
            log.info(f"Cloudflare resolved after ~{(i+1)*2}s")
            time.sleep(2)
            return
        if i % 15 == 14:
            log.info(f"Still waiting on Cloudflare... ({(i+1)*2}s)")
    raise TimeoutError("Cloudflare challenge not resolved.")


def login(driver):
    log.info("Navigating to Steel City Vacuum...")
    driver.get(CONFIG["base_url"])
    random_delay()
    wait_for_cloudflare(driver)

    try:
        nav = driver.find_element(By.ID, "main-nav")
        if "notlogged" not in (nav.get_attribute("class") or "").lower():
            log.info("Already logged in!")
            return
    except Exception:
        pass

    log.info("Logging in...")
    try:
        driver.find_element(By.CSS_SELECTOR, "a.nav-login-btn").click()
        time.sleep(1)
    except Exception:
        pass

    wait = WebDriverWait(driver, 10)
    customer_field = wait.until(EC.visibility_of_element_located((By.ID, "scvCustomerNumber")))
    human_type(customer_field, CONFIG["account"])
    random_delay(0.5, 1.0)
    human_type(driver.find_element(By.ID, "userNameBox"), CONFIG["user_id"])
    random_delay(0.5, 1.0)
    human_type(driver.find_element(By.ID, "password"), CONFIG["password"])
    random_delay(0.5, 1.0)

    try:
        driver.find_element(By.CSS_SELECTOR, "a.login-btn").click()
    except Exception:
        driver.execute_script("submitLoginForm()")

    time.sleep(5)
    wait_for_cloudflare(driver)
    log.info("Login complete.")


# ── API Calls (via browser JS) ─────────────────────────────────────────────

def api_search(driver, query, take=200):
    """Call the search API. Returns list of result objects."""
    try:
        result = driver.execute_script("""
            return new Promise((resolve) => {
                $.ajax({
                    type: 'POST',
                    url: 'applications/shopping_cart/web_services.php?action=search&take=' + arguments[1] + '&searchstring=' + encodeURIComponent(arguments[0]),
                    success: function(data) {
                        if (typeof data === 'string') {
                            try { data = JSON.parse(data); } catch(e) {}
                        }
                        resolve(JSON.stringify(data));
                    },
                    error: function() { resolve(null); }
                });
            });
        """, query, take)
        if result:
            return json.loads(result)
    except Exception as e:
        log.debug(f"Search API failed for '{query}': {e}")
    return None


def api_autocomplete(driver, query, max_rows=100):
    """Call the autocomplete API. Returns list of result objects."""
    try:
        result = driver.execute_script("""
            return new Promise((resolve) => {
                $.ajax({
                    type: 'POST',
                    url: 'applications/shopping_cart/web_services.php?action=item_autocomplete',
                    data: { name: arguments[0], maxRows: arguments[1] },
                    success: function(data) {
                        if (typeof data === 'string') {
                            try { data = JSON.parse(data); } catch(e) {}
                        }
                        resolve(JSON.stringify(data));
                    },
                    error: function() { resolve(null); }
                });
            });
        """, query, max_rows)
        if result:
            return json.loads(result)
    except Exception as e:
        log.debug(f"Autocomplete API failed for '{query}': {e}")
    return None


def api_product_info(driver, part_id):
    """Call the product_info API. Returns product dict or None."""
    try:
        result = driver.execute_script("""
            return new Promise((resolve) => {
                $.ajax({
                    type: 'POST',
                    url: 'applications/shopping_cart/web_services.php?action=product_info&name=' + arguments[0],
                    success: function(data) {
                        if (typeof data === 'string') {
                            try { data = JSON.parse(data); } catch(e) {}
                        }
                        resolve(JSON.stringify(data));
                    },
                    error: function() { resolve(null); }
                });
            });
        """, part_id)
        if result:
            return json.loads(result)
    except Exception as e:
        log.debug(f"Product info API failed for '{part_id}': {e}")
    return None


def api_category_products(driver, category_id, page=1, take=50):
    """Try to get products from a category via AJAX. Returns list or None."""
    try:
        result = driver.execute_script("""
            return new Promise((resolve) => {
                $.ajax({
                    type: 'POST',
                    url: 'applications/shopping_cart/web_services.php?action=search&take=' + arguments[2] + '&category=' + arguments[0] + '&page=' + arguments[1],
                    success: function(data) {
                        if (typeof data === 'string') {
                            try { data = JSON.parse(data); } catch(e) {}
                        }
                        resolve(JSON.stringify(data));
                    },
                    error: function() { resolve(null); }
                });
            });
        """, category_id, page, take)
        if result:
            return json.loads(result)
    except Exception as e:
        log.debug(f"Category API failed for cat {category_id}: {e}")
    return None


# ── Phase 1: Discovery ─────────────────────────────────────────────────────

def discover_categories(driver, progress):
    """Discover product categories from the store section."""
    log.info("=" * 60)
    log.info("DISCOVERING CATEGORIES")
    log.info("=" * 60)

    # Navigate to store page
    driver.get(CONFIG["store_url"])
    random_delay(2, 4)
    wait_for_cloudflare(driver)
    save_html(driver, "store_main")
    screenshot(driver, "store_main")

    categories = {}

    # Method 1: Extract category links from the store page
    log.info("Looking for category links on store page...")
    try:
        # Look for links matching /a/s/c/{id} pattern
        links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/a/s/c/']")
        for link in links:
            href = link.get_attribute("href") or ""
            name = link.text.strip()
            match = re.search(r'/a/s/c/(\d+)', href)
            if match and name:
                cat_id = match.group(1)
                categories[cat_id] = name
                log.info(f"  Found category: {name} (ID: {cat_id})")
    except Exception as e:
        log.warning(f"Error extracting category links: {e}")

    # Method 2: Use autocomplete to discover categories
    log.info("Searching for categories via autocomplete API...")
    cat_search_terms = [
        "bag", "filter", "belt", "brush", "hose", "cord", "motor",
        "cleaner", "vacuum", "attachment", "tool", "hepa", "central",
        "parts", "accessories", "supply", "commercial", "residential",
    ]
    for term in cat_search_terms:
        results = api_autocomplete(driver, term, max_rows=50)
        if results and isinstance(results, list):
            for r in results:
                if isinstance(r, dict) and r.get("type") == "category":
                    cat_id = str(r.get("categoryID", ""))
                    cat_name = r.get("name", "")
                    if cat_id and cat_name and cat_id not in categories:
                        categories[cat_id] = cat_name
                        log.info(f"  Found category via autocomplete: {cat_name} (ID: {cat_id})")
        time.sleep(random.uniform(0.3, 0.6))

    # Category IDs go up to 12,000+ — brute-force is impractical.
    # The autocomplete + page link approach above should find the main categories.
    log.info(f"  Skipping category ID brute-force (IDs range into thousands).")

    progress["categories_discovered"] = [
        {"id": cid, "name": cname} for cid, cname in categories.items()
    ]
    save_progress(progress)
    log.info(f"Category discovery complete. Found {len(categories)} categories.")
    return categories


def discover_products_via_search(driver, existing_parts, progress):
    """Use the search API with broad terms to find products not in schematics."""
    log.info("=" * 60)
    log.info("DISCOVERING PRODUCTS VIA SEARCH API")
    log.info("=" * 60)

    done_terms = set(progress.get("search_terms_done", []))
    remaining_terms = [t for t in SEARCH_TERMS if t not in done_terms]
    log.info(f"Search terms: {len(remaining_terms)} remaining (of {len(SEARCH_TERMS)} total)")

    new_found = 0

    for i, term in enumerate(remaining_terms):
        # Try search API
        # Response format: {"results": [...], "total": N, "cat_brands": [...], "cat_cats": [...]}
        # Each result is a dict with numeric string keys:
        #   "0": internal_id, "1": productID, "2": product_code/SKU, "3": name/description,
        #   "4": ?, "5": price, "6": ?, "7": in_stock, "8": ?, ...
        results = api_search(driver, term, take=500)

        if results and isinstance(results, dict):
            search_results = results.get("results", [])
            total = results.get("total", 0)
            if i < 3:
                log.info(f"  Search '{term}': {len(search_results)} results (total: {total})")

            for item in search_results:
                if not isinstance(item, dict):
                    continue

                product_code = str(item.get("2", "")).strip()
                product_id = str(item.get("1", "")).strip()
                raw_name = str(item.get("3", "")).strip()
                price = str(item.get("5", "")).strip()

                if not product_code:
                    continue

                # Skip existing parts
                if product_code in existing_parts:
                    continue
                if product_code in progress["discovered_products"]:
                    continue

                progress["discovered_products"][product_code] = {
                    "source": f"search:{term}",
                    "product_id": product_id,
                    "raw_name": raw_name,
                    "raw_product_code": product_code,
                    "raw_price": price,
                    "manufacturer": "",
                }
                new_found += 1

        # Also try autocomplete for product-type results
        # Autocomplete returns: {"name": "...", "type": "product"|"category"|..., "productID": N, ...}
        auto_results = api_autocomplete(driver, term, max_rows=100)
        if auto_results and isinstance(auto_results, list):
            for r in auto_results:
                if not isinstance(r, dict):
                    continue
                if r.get("type") != "product":
                    continue
                # Autocomplete product results have 'name' (display name) and 'productID'
                # but NOT product_code — we need to enrich these later via product page
                product_id = str(r.get("productID", ""))
                if not product_id or product_id == "-2":
                    continue
                key = f"pid_{product_id}"
                if key in progress["discovered_products"]:
                    continue
                progress["discovered_products"][key] = {
                    "source": f"autocomplete:{term}",
                    "product_id": product_id,
                    "raw_name": r.get("name", ""),
                    "raw_product_code": "",
                    "manufacturer": r.get("manufacturer", ""),
                }
                new_found += 1

        progress["search_terms_done"].append(term)
        time.sleep(random.uniform(0.3, 0.8))

        if (i + 1) % 10 == 0 or i < 3:
            save_progress(progress)
            log.info(f"  [{i+1}/{len(remaining_terms)}] Search terms done, {new_found} new products found (total discovered: {len(progress['discovered_products'])})")

    save_progress(progress)
    log.info(f"Search discovery complete. {new_found} new products found this run.")
    log.info(f"Total discovered products: {len(progress['discovered_products'])}")


def scrape_category_pages(driver, categories, existing_parts, progress):
    """Use search API with category filter to get products from each category."""
    log.info("=" * 60)
    log.info("SCRAPING CATEGORY PAGES VIA SEARCH API")
    log.info("=" * 60)

    scraped = set(progress.get("categories_scraped", []))
    remaining = [(cid, cname) for cid, cname in categories.items() if cid not in scraped]
    log.info(f"Categories to scrape: {len(remaining)} remaining")

    new_found = 0

    for idx, (cat_id, cat_name) in enumerate(remaining):
        # Navigate to the category page — this also lets the search API filter by category context
        try:
            driver.get(f"{CONFIG['base_url']}/a/s/c/{cat_id}")
            time.sleep(2)
            wait_for_cloudflare(driver)

            # Extract product links from the rendered page
            product_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/a/s/pid/']")
            found_on_page = 0

            for link in product_links:
                href = link.get_attribute("href") or ""
                name = link.text.strip()
                pid_match = re.search(r'/a/s/pid/(\d+)', href)
                product_id = pid_match.group(1) if pid_match else ""

                if product_id:
                    key = f"pid_{product_id}"
                    if key not in progress["discovered_products"] and key not in existing_parts:
                        progress["discovered_products"][key] = {
                            "source": f"category:{cat_id}:{cat_name}",
                            "product_id": product_id,
                            "raw_name": name,
                            "raw_product_code": "",
                            "category": cat_name,
                        }
                        new_found += 1
                        found_on_page += 1

            # Check for pagination — get subsequent pages
            page = 1
            while True:
                try:
                    next_link = driver.find_element(By.LINK_TEXT, "Next")
                    next_url = next_link.get_attribute("href")
                    if next_url:
                        page += 1
                        driver.get(next_url)
                        time.sleep(1.5)
                        wait_for_cloudflare(driver)
                        more_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/a/s/pid/']")
                        for link in more_links:
                            href = link.get_attribute("href") or ""
                            name = link.text.strip()
                            pid_match = re.search(r'/a/s/pid/(\d+)', href)
                            product_id = pid_match.group(1) if pid_match else ""
                            if product_id:
                                key = f"pid_{product_id}"
                                if key not in progress["discovered_products"] and key not in existing_parts:
                                    progress["discovered_products"][key] = {
                                        "source": f"category:{cat_id}:{cat_name}",
                                        "product_id": product_id,
                                        "raw_name": name,
                                        "raw_product_code": "",
                                        "category": cat_name,
                                    }
                                    new_found += 1
                                    found_on_page += 1
                    else:
                        break
                except NoSuchElementException:
                    break

            if found_on_page > 0 or idx < 5:
                log.info(f"  Category '{cat_name}' (ID: {cat_id}): {found_on_page} new products (pages: {page})")

        except Exception as e:
            log.warning(f"  Error on category {cat_id}: {e}")

        progress["categories_scraped"].append(cat_id)
        if (idx + 1) % 10 == 0:
            save_progress(progress)
            log.info(f"  [{idx+1}/{len(remaining)}] Categories scraped, {new_found} new products total")

    save_progress(progress)
    log.info(f"Category scraping complete. {new_found} new products found.")


# ── Phase 2: Enrichment ────────────────────────────────────────────────────

def enrich_products(driver, existing_parts, progress):
    """Call product_info API for each discovered product to get full details."""
    log.info("=" * 60)
    log.info("ENRICHING PRODUCTS VIA PRODUCT_INFO API")
    log.info("=" * 60)

    already_enriched = set(progress.get("enriched_products", {}).keys())
    to_enrich = [
        (key, info) for key, info in progress["discovered_products"].items()
        if key not in already_enriched
    ]
    log.info(f"Products to enrich: {len(to_enrich)} remaining")

    enriched = 0
    skipped_existing = 0
    skipped_special = 0
    skipped_nla = 0

    for i, (key, info) in enumerate(to_enrich):
        product_id = info.get("product_id", "")
        raw_code = info.get("raw_product_code", "")

        # Determine what to query the API with
        # If key doesn't start with "pid_", it IS the product_code (from search API)
        # If it starts with "pid_", we have a productID and need to resolve the part number
        if not key.startswith("pid_"):
            part_id = key
        elif raw_code:
            part_id = raw_code
        else:
            # Need to visit product page to get the actual part number
            part_id = ""
            if product_id:
                try:
                    driver.get(f"{CONFIG['base_url']}/a/s/pid/{product_id}")
                    time.sleep(1.5)
                    wait_for_cloudflare(driver)

                    # Try to extract product code from the page
                    page_source = driver.page_source
                    code_match = re.search(r'product_code["\s:]+["\']?([A-Za-z0-9\-_.]+)', page_source)
                    if code_match:
                        part_id = code_match.group(1)
                    else:
                        try:
                            code_el = driver.find_element(By.CSS_SELECTOR, ".product-code, .sku, [class*='product-code'], [class*='sku']")
                            part_id = code_el.text.strip()
                        except NoSuchElementException:
                            pass
                except Exception as e:
                    log.debug(f"Error visiting product page for {key}: {e}")

        if not part_id:
            log.debug(f"  Could not determine part ID for {key}, skipping")
            continue

        # Skip if this part number is in existing schematics data
        if part_id in existing_parts:
            skipped_existing += 1
            continue

        # Call product_info API
        data = api_product_info(driver, part_id)
        time.sleep(random.uniform(0.2, 0.5))

        if not data or not data.get("name"):
            log.debug(f"  No data for {part_id}")
            continue

        # Extract fields
        name = data.get("name", "")
        product_code = data.get("product_code", part_id)
        description = data.get("description", "")
        price = data.get("Price_1", "")
        in_stock = data.get("in_stock", "")
        picture = data.get("picture", "")
        big_picture = data.get("big_picture", "")
        alt_items = data.get("alt_items", [])
        manufacturer = data.get("manufacturer", info.get("manufacturer", ""))

        # Check if this product_code is already known
        if product_code in existing_parts:
            skipped_existing += 1
            continue

        # Check for NLA
        is_nla = False
        if description:
            desc_upper = description.upper()
            if "NLA" in desc_upper or "NO LONGER AVAILABLE" in desc_upper:
                is_nla = True
        if name:
            name_upper = name.upper()
            if "NLA" in name_upper or "NO LONGER AVAILABLE" in name_upper:
                is_nla = True

        # Parse alt items
        alt_strs = []
        if alt_items and isinstance(alt_items, list):
            for alt in alt_items:
                if isinstance(alt, dict) and alt.get("name"):
                    alt_str = alt["name"]
                    if alt.get("product_code"):
                        alt_str += f" ({alt['product_code']})"
                    alt_strs.append(alt_str)

        # Filter: NLA without alternatives → skip
        if is_nla and not alt_strs:
            skipped_nla += 1
            continue

        # Check for special order (no image AND certain markers)
        is_special_order = False
        if description and "SPECIAL ORDER" in description.upper():
            is_special_order = True
        if not picture or picture == "0":
            # No image — could be special order
            if "SPECIAL" in (description or "").upper() or "SPECIAL" in (name or "").upper():
                is_special_order = True

        if is_special_order:
            skipped_special += 1
            continue

        # Build image URL
        image_url = ""
        img_filename = (big_picture or picture or "").strip()
        if img_filename and img_filename != "0":
            image_url = IMAGE_BASE_URL + img_filename

        # Format price
        price_str = ""
        if price:
            try:
                price_str = f"${float(price):.2f}"
            except (ValueError, TypeError):
                price_str = str(price)

        # Volume pricing
        vol_prices = []
        for qty in ["5", "10", "25", "50"]:
            p = data.get(f"Price_{qty}")
            if p:
                try:
                    vol_prices.append(f"{qty}+: ${float(p):.2f}")
                except (ValueError, TypeError):
                    pass

        notes_parts = []
        if vol_prices:
            notes_parts.append(" | ".join(vol_prices))
        if alt_strs:
            notes_parts.append("ALT: " + "; ".join(alt_strs))
        if is_nla:
            notes_parts.append("NLA (has alternative)")

        # Store enriched product
        progress["enriched_products"][product_code] = {
            "part_number": part_id,
            "product_code": product_code,
            "name": name,
            "description": description,
            "price": price_str,
            "in_stock": in_stock,
            "image_url": image_url,
            "picture": picture,
            "big_picture": big_picture,
            "alt_items": "; ".join(alt_strs) if alt_strs else "",
            "manufacturer": manufacturer,
            "category": info.get("category", ""),
            "source": info.get("source", ""),
            "notes": " | ".join(notes_parts),
            "product_id": str(data.get("productID", product_id)),
        }
        enriched += 1

        if (i + 1) % CHECKPOINT_INTERVAL == 0:
            progress["skipped_existing"] = skipped_existing
            progress["skipped_special_order"] = skipped_special
            progress["skipped_nla"] = skipped_nla
            save_progress(progress)
            log.info(f"  [{i+1}/{len(to_enrich)}] Enriched: {enriched}, "
                     f"Skipped existing: {skipped_existing}, "
                     f"Skipped special order: {skipped_special}, "
                     f"Skipped NLA: {skipped_nla}")

    progress["skipped_existing"] = skipped_existing
    progress["skipped_special_order"] = skipped_special
    progress["skipped_nla"] = skipped_nla
    save_progress(progress)
    log.info(f"Enrichment complete. Enriched: {enriched}")
    log.info(f"  Skipped existing: {skipped_existing}")
    log.info(f"  Skipped special order: {skipped_special}")
    log.info(f"  Skipped NLA (no alt): {skipped_nla}")


# ── Phase 3: Export ─────────────────────────────────────────────────────────

def export_to_excel(progress):
    """Export enriched products to Excel."""
    log.info("=" * 60)
    log.info("EXPORTING TO EXCEL")
    log.info("=" * 60)

    products = progress.get("enriched_products", {})
    if not products:
        log.warning("No products to export!")
        return

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "New Catalog Products"

    headers = [
        "Part Number", "SKU", "Name", "Description", "Price",
        "In Stock", "Image URL", "Alt Items", "Manufacturer",
        "Category", "Source", "Notes", "Product ID",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for code, product in products.items():
        ws.append([
            product.get("part_number", ""),
            product.get("product_code", ""),
            product.get("name", ""),
            product.get("description", ""),
            product.get("price", ""),
            product.get("in_stock", ""),
            product.get("image_url", ""),
            product.get("alt_items", ""),
            product.get("manufacturer", ""),
            product.get("category", ""),
            product.get("source", ""),
            product.get("notes", ""),
            product.get("product_id", ""),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"

    filepath = OUTPUT_DIR / "catalog_new_products.xlsx"
    wb.save(filepath)
    log.info(f"Excel saved: {filepath} ({len(products)} products)")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Steel City Catalog Scraper")
    parser.add_argument("--step", choices=["discover", "enrich", "export", "all"], default="all",
                        help="Which step to run")
    args = parser.parse_args()

    progress = load_progress()
    existing_parts = load_existing_part_numbers()

    needs_browser = args.step in ("discover", "enrich", "all")
    driver = None

    try:
        if needs_browser:
            driver = create_driver()
            login(driver)

            # Navigate to a page on the site so AJAX calls work (same-origin)
            driver.get(CONFIG["store_url"])
            time.sleep(3)
            wait_for_cloudflare(driver)

        if args.step in ("discover", "all"):
            # Step 1: Discover categories
            categories = discover_categories(driver, progress)

            # Step 2: Scrape category pages for product listings
            scrape_category_pages(driver, categories, existing_parts, progress)

            # Step 3: Search API discovery
            discover_products_via_search(driver, existing_parts, progress)

        if args.step in ("enrich", "all"):
            # Step 4: Enrich all discovered products
            enrich_products(driver, existing_parts, progress)

        if args.step in ("export", "all"):
            # Step 5: Export to Excel
            export_to_excel(progress)

    except KeyboardInterrupt:
        log.info("Interrupted — progress saved.")
        save_progress(progress)
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    # Print summary
    log.info("=" * 60)
    log.info("SUMMARY")
    log.info(f"  Existing parts (schematics):  {len(existing_parts)}")
    log.info(f"  Categories discovered:        {len(progress.get('categories_discovered', []))}")
    log.info(f"  Products discovered (raw):    {len(progress.get('discovered_products', {}))}")
    log.info(f"  Products enriched (new):      {len(progress.get('enriched_products', {}))}")
    log.info(f"  Skipped (already had):        {progress.get('skipped_existing', 0)}")
    log.info(f"  Skipped (special order):      {progress.get('skipped_special_order', 0)}")
    log.info(f"  Skipped (NLA, no alt):        {progress.get('skipped_nla', 0)}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
