#!/usr/bin/env python3
"""
Replace flagged product images by searching Google Images for better alternatives.

Reads visual_review_results.json, looks up brand/description/SKU for each part,
searches Google Images using product description + brand + SKU, downloads the first
result, and processes it through the same pipeline (rembg → VaM logo → pad 2048x2048).

Saves as 1_google.jpg alongside the original 1.jpg for manual comparison.
"""

import json
import os
import re
import time
import urllib.parse

import numpy as np
import requests
from PIL import Image
from rembg import remove

# ─── Configuration ─────────────────────────────────────────────────────────────

FLAGGED_JSON = "visual_review_results.json"
PROGRESS_JSON = "replace_progress.json"
IMAGES_DIR = "images"
IMAGES_RAW_DIR = "images_raw"
LOGO_PATH = "VaM Watermark.png"
EXCEL_SCHEMATIC = "output/steel_city_parts.xlsx"
EXCEL_CATALOG = "output/catalog_new_products.xlsx"

CANVAS_SIZE = 2048
PRODUCT_MAX = 1800
LOGO_SCALE = 0.80
LOGO_OPACITY = 0.20

SEARCH_DELAY = 2.0  # seconds between Google searches to avoid rate limiting

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


# ─── Part Data Lookup ──────────────────────────────────────────────────────────

def load_part_data():
    """Load brand, description, SKU, model for all flagged parts from Excel files."""
    import openpyxl

    with open(FLAGGED_JSON) as f:
        data = json.load(f)

    all_parts = {}
    for item in data.get("watermark_visible", []):
        all_parts[item["part"]] = {"issue": "watermark"}
    for item in data.get("background_removal_failure", []):
        all_parts[item["part"]] = {"issue": "bg_removal"}

    # Schematics Excel
    wb = openpyxl.load_workbook(EXCEL_SCHEMATIC, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        pn = str(row_dict.get("Part Number", "")).strip()
        if pn in all_parts and "brand" not in all_parts[pn]:
            brand = str(row_dict.get("Brand", "") or "").strip()
            if brand:
                all_parts[pn]["brand"] = brand
                all_parts[pn]["name"] = str(row_dict.get("Name", "") or "").strip()
                all_parts[pn]["description"] = str(row_dict.get("Description", "") or "").strip()
                all_parts[pn]["sku"] = str(row_dict.get("SKU", "") or "").strip()
                all_parts[pn]["model"] = str(row_dict.get("Model", "") or "").strip()
    wb.close()

    # Catalog Excel for remaining
    missing = [pn for pn in all_parts if "brand" not in all_parts[pn]]
    if missing:
        wb2 = openpyxl.load_workbook(EXCEL_CATALOG, read_only=True)
        ws2 = wb2.active
        headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers2, row))
            pn = str(row_dict.get("Part Number", "")).strip()
            if pn in all_parts and "brand" not in all_parts[pn]:
                brand = str(row_dict.get("Manufacturer", "") or "").strip()
                if brand:
                    all_parts[pn]["brand"] = brand
                    all_parts[pn]["name"] = str(row_dict.get("Name", "") or "").strip()
                    all_parts[pn]["description"] = str(row_dict.get("Description", "") or "").strip()
                    all_parts[pn]["sku"] = str(row_dict.get("SKU", "") or "").strip()
                    all_parts[pn]["model"] = ""
        wb2.close()

    # Defaults for anything still missing
    for pn in all_parts:
        all_parts[pn].setdefault("brand", "")
        all_parts[pn].setdefault("name", "")
        all_parts[pn].setdefault("description", "")
        all_parts[pn].setdefault("sku", "")
        all_parts[pn].setdefault("model", "")

    return all_parts


def build_search_query(pn, info):
    """
    Build an optimal Google Images search query from part data.

    The Description field often has the best product name, e.g.:
      "01-7887-01 - KNOB-BERNINA 6100 POST FILTER COVER/BLACK"
      "21-8605-05 - FAN-LEXAN-OEM EUREKA 2000 SERIES"

    Strategy: extract the product description (after the " - "), combine with
    brand and SKU for the most specific query possible.
    """
    brand = info["brand"].split(",")[0].strip() if info["brand"] else ""
    desc = info["description"]
    sku = info["sku"]
    name = info["name"]

    # Extract the descriptive part from Description field
    # Format is typically "PARTNUM - ACTUAL DESCRIPTION"
    product_desc = ""
    if desc:
        if " - " in desc:
            product_desc = desc.split(" - ", 1)[1].strip()
        else:
            product_desc = desc.strip()

    # Clean up the description: remove "DO NOT USE" type warnings
    if product_desc:
        product_desc = re.sub(r'\bDO NOT USE.*', '', product_desc, flags=re.IGNORECASE).strip()
        # Replace hyphens and slashes with spaces for better search
        product_desc = product_desc.replace("-", " ").replace("/", " ")
        # Remove OEM references that clutter search
        product_desc = re.sub(r'\bOEM\b', '', product_desc, flags=re.IGNORECASE).strip()

    # Build query: brand + product description + SKU
    parts = []
    if brand:
        parts.append(brand)
    if product_desc:
        parts.append(product_desc)
    elif name:
        parts.append(name)

    # Add SKU/part number for specificity
    if sku and sku != pn:
        parts.append(sku)
    else:
        parts.append(pn)

    query = " ".join(parts)

    # If query is too short (no description found), fall back to basic search
    if len(query) < 10:
        query = f"{brand} {pn} vacuum part".strip()

    return query


# ─── Google Image Search ───────────────────────────────────────────────────────

def search_google_images(query, num_results=5):
    """
    Search Google Images and return URLs of the first few results.
    """
    encoded_query = urllib.parse.quote(query)
    url = f"https://www.google.com/search?q={encoded_query}&tbm=isch&safe=active"

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        html = resp.text

        img_urls = []
        patterns = [
            r'"(https?://[^"]+\.(?:jpg|jpeg|png|webp)(?:\?[^"]*)?)"',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, html, re.IGNORECASE)
            for match in matches:
                if any(skip in match.lower() for skip in [
                    "google.com", "gstatic.com", "googleapis.com",
                    "youtube.com", "ytimg.com", "favicon",
                    "logo", "icon", "badge", "sprite",
                    "encrypted-tbn", "x-raw-image",
                ]):
                    continue
                if match not in img_urls:
                    img_urls.append(match)
                    if len(img_urls) >= num_results:
                        break
            if len(img_urls) >= num_results:
                break

        return img_urls

    except Exception as e:
        print(f"    Search error: {e}")
        return []


def download_image(url, save_path, timeout=15):
    """Download an image from URL and save it."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
        resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "")
        if "image" not in content_type and not url.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
            return False

        with open(save_path, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)

        # Verify valid image
        img = Image.open(save_path)
        img.verify()

        # Check minimum size
        img = Image.open(save_path)
        if img.width < 100 or img.height < 100:
            os.remove(save_path)
            return False

        return True

    except Exception:
        if os.path.exists(save_path):
            os.remove(save_path)
        return False


# ─── Image Processing Pipeline ─────────────────────────────────────────────────

def load_logo():
    """Load and prepare the VaM watermark logo."""
    logo = Image.open(LOGO_PATH).convert("RGBA")
    target = int(CANVAS_SIZE * LOGO_SCALE)
    logo.thumbnail((target, target), Image.LANCZOS)

    r, g, b, a = logo.split()
    a = a.point(lambda x: int(x * LOGO_OPACITY))
    logo.putalpha(a)
    return logo


def process_image(raw_path, output_path, logo):
    """Process a raw image: remove bg, composite with logo, pad to canvas."""
    raw_img = Image.open(raw_path).convert("RGBA")
    nobg = remove(raw_img)

    bbox = nobg.getbbox()
    if not bbox:
        return False

    product = nobg.crop(bbox)
    product.thumbnail((PRODUCT_MAX, PRODUCT_MAX), Image.LANCZOS)

    canvas = Image.new("RGBA", (CANVAS_SIZE, CANVAS_SIZE), (255, 255, 255, 255))

    lx = (CANVAS_SIZE - logo.width) // 2
    ly = (CANVAS_SIZE - logo.height) // 2
    canvas.paste(logo, (lx, ly), logo)

    px = (CANVAS_SIZE - product.width) // 2
    py = (CANVAS_SIZE - product.height) // 2
    canvas.paste(product, (px, py), product)

    final = canvas.convert("RGB")
    final.save(output_path, "JPEG", quality=95)
    return True


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading flagged parts and product data...")
    parts = load_part_data()
    print(f"Total flagged parts: {len(parts)}")

    # Load progress
    progress = {}
    if os.path.exists(PROGRESS_JSON):
        with open(PROGRESS_JSON) as f:
            progress = json.load(f)

    logo = load_logo()

    success = 0
    failed = 0
    skipped = 0

    for i, (pn, info) in enumerate(sorted(parts.items())):
        if progress.get(pn, {}).get("status") == "done":
            skipped += 1
            continue

        query = build_search_query(pn, info)
        brand = info["brand"].split(",")[0] if info["brand"] else "unknown"

        print(f"\n[{i+1}/{len(parts)}] {pn} (brand: {brand})")
        print(f"  Query: {query}")

        # Search Google Images
        img_urls = search_google_images(query)

        if not img_urls:
            print(f"  No images found!")
            progress[pn] = {"status": "no_results", "query": query}
            failed += 1
            _save_progress(progress)
            time.sleep(SEARCH_DELAY)
            continue

        # Try downloading each result until one works
        raw_path = os.path.join(IMAGES_RAW_DIR, f"{pn}_google.jpg")
        output_dir = os.path.join(IMAGES_DIR, pn)
        # Save as 1_google.jpg — don't overwrite original
        output_path = os.path.join(output_dir, "1_google.jpg")

        downloaded = False
        download_url = None
        for j, url in enumerate(img_urls):
            print(f"  Trying URL {j+1}/{len(img_urls)}...")
            if download_image(url, raw_path):
                downloaded = True
                download_url = url
                print(f"  Downloaded!")
                break

        if not downloaded:
            print(f"  All download attempts failed!")
            progress[pn] = {"status": "download_failed", "query": query, "urls_tried": len(img_urls)}
            failed += 1
            _save_progress(progress)
            time.sleep(SEARCH_DELAY)
            continue

        # Process through pipeline
        os.makedirs(output_dir, exist_ok=True)
        try:
            if process_image(raw_path, output_path, logo):
                print(f"  Processed → {output_path}")
                progress[pn] = {"status": "done", "query": query, "source_url": download_url}
                success += 1
            else:
                print(f"  Processing failed (no product detected)!")
                progress[pn] = {"status": "process_failed", "query": query}
                failed += 1
        except Exception as e:
            print(f"  Processing error: {e}")
            progress[pn] = {"status": "process_error", "query": query, "error": str(e)}
            failed += 1

        _save_progress(progress)
        time.sleep(SEARCH_DELAY)

    print(f"\n{'='*50}")
    print(f"Done!")
    print(f"  Success:  {success}")
    print(f"  Failed:   {failed}")
    print(f"  Skipped:  {skipped} (already done)")
    print(f"  Total:    {len(parts)}")


def _save_progress(progress):
    with open(PROGRESS_JSON, "w") as f:
        json.dump(progress, f, indent=2)


if __name__ == "__main__":
    main()
