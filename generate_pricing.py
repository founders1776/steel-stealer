#!/usr/bin/env python3
"""
Generate competitive retail prices for Steel Stealer Shopify products.

Applies tiered markup to dealer cost, charm pricing (.99), and minimum floor.
Updates product_names.json and output/product_descriptions.xlsx.
"""

import json
import math
import os

import openpyxl

# ── Pricing Configuration ────────────────────────────────────────────────────

# (max_cost, multiplier) — sorted ascending by max_cost
MARKUP_TIERS = [
    (1.00,    8.0),   # Tiny parts: belts, screws, springs
    (3.00,    4.5),   # Small parts, bags, basic filters
    (7.00,    3.2),   # Bulk of cheap catalog items
    (15.00,   2.5),   # Mid-range filters, cords, belts
    (30.00,   2.2),   # Brush rolls, hoses, assemblies
    (60.00,   1.9),   # Power nozzles, quality filters
    (120.00,  1.7),   # Motors, power heads
    (300.00,  1.5),   # Premium motors, full assemblies
    (float("inf"), 1.4),  # Big-ticket items
]

MIN_PRICE = 6.99  # No product listed below this


# ── Pricing Logic ─────────────────────────────────────────────────────────────

def get_markup(cost):
    """Return the markup multiplier for a given dealer cost."""
    for max_cost, multiplier in MARKUP_TIERS:
        if cost <= max_cost:
            return multiplier
    return MARKUP_TIERS[-1][1]


def charm_price(raw_price):
    """Round to nearest .99 using charm pricing.

    If raw is e.g. $12.40 → $11.99 (round down)
    If raw is e.g. $12.80 → $12.99 (round up)
    Threshold: .50 of the dollar — below rounds down, at/above rounds up.
    """
    dollar = int(raw_price)
    cents = raw_price - dollar

    if cents < 0.50:
        # Round down to previous .99
        return float(dollar - 1) + 0.99 if dollar > 0 else 0.99
    else:
        # Round up to this .99
        return float(dollar) + 0.99


def calculate_retail_price(cost):
    """Apply tiered markup + charm pricing + minimum floor."""
    markup = get_markup(cost)
    raw = cost * markup
    retail = charm_price(raw)
    return max(retail, MIN_PRICE)


def parse_price(price_str):
    """Extract numeric price from string like '$11.92'."""
    if not price_str:
        return None
    import re
    match = re.search(r'[\d]+\.?\d*', str(price_str))
    return float(match.group()) if match else None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Load product data
    with open("product_names.json", "r") as f:
        products = json.load(f)

    print(f"Loaded {len(products)} products from product_names.json")

    # Apply pricing
    priced = 0
    skipped = 0
    prices = []
    costs = []
    margins = []

    for key, product in products.items():
        cost = parse_price(product.get("price"))
        if cost is None or cost <= 0:
            skipped += 1
            continue

        retail = calculate_retail_price(cost)
        product["retail_price"] = f"${retail:.2f}"
        priced += 1
        prices.append(retail)
        costs.append(cost)
        margins.append((retail - cost) / retail * 100)

    print(f"\nPriced: {priced} | Skipped (no cost): {skipped}")

    # Save updated product_names.json
    with open("product_names.json", "w") as f:
        json.dump(products, f, indent=2)
    print("Updated product_names.json with retail_price field")

    # Update Excel spreadsheet
    update_spreadsheet(products)

    # Print summary
    print_summary(prices, costs, margins)


def update_spreadsheet(products):
    """Add Retail Price column to output/product_descriptions.xlsx."""
    xlsx_path = "output/product_descriptions.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"\nWARNING: {xlsx_path} not found — skipping Excel update")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find headers
    headers = [cell.value for cell in ws[1]]

    # Find SKU column (to match products)
    sku_col = None
    for i, h in enumerate(headers):
        if h and "sku" in str(h).lower():
            sku_col = i
            break

    if sku_col is None:
        print("WARNING: No SKU column found in spreadsheet — skipping Excel update")
        return

    # Check if Retail Price column already exists
    retail_col = None
    for i, h in enumerate(headers):
        if h and "retail" in str(h).lower() and "price" in str(h).lower():
            retail_col = i + 1  # 1-indexed for openpyxl
            break

    if retail_col is None:
        # Add new column at the end
        retail_col = len(headers) + 1
        ws.cell(row=1, column=retail_col, value="Retail Price")
        print(f"Added 'Retail Price' column at position {retail_col}")

    # Build lookup: try matching by SKU or by product key
    product_by_sku = {}
    for key, p in products.items():
        if "retail_price" in p:
            product_by_sku[p.get("sku", "")] = p["retail_price"]
            product_by_sku[key] = p["retail_price"]

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        sku_val = str(ws.cell(row=row_idx, column=sku_col + 1).value or "").strip()
        retail = product_by_sku.get(sku_val)
        if retail:
            ws.cell(row=row_idx, column=retail_col, value=retail)
            updated += 1

    wb.save(xlsx_path)
    print(f"Updated {xlsx_path}: {updated} rows with retail prices")


def print_summary(prices, costs, margins):
    """Print pricing summary statistics."""
    if not prices:
        print("No prices to summarize.")
        return

    print("\n" + "=" * 60)
    print("PRICING SUMMARY")
    print("=" * 60)

    # Overall stats
    avg_margin = sum(margins) / len(margins)
    avg_retail = sum(prices) / len(prices)
    avg_cost = sum(costs) / len(costs)
    median_retail = sorted(prices)[len(prices) // 2]

    print(f"\nAvg dealer cost:  ${avg_cost:.2f}")
    print(f"Avg retail price: ${avg_retail:.2f}")
    print(f"Median retail:    ${median_retail:.2f}")
    print(f"Avg margin:       {avg_margin:.1f}%")

    # Distribution
    buckets = [
        (0, 10), (10, 25), (25, 50), (50, 100),
        (100, 250), (250, 500), (500, float("inf")),
    ]
    print(f"\n{'Retail Range':<20} {'Count':>6} {'%':>6}")
    print("-" * 34)
    for lo, hi in buckets:
        count = sum(1 for p in prices if lo <= p < hi)
        pct = count / len(prices) * 100
        label = f"${lo}–${hi}" if hi != float("inf") else f"${lo}+"
        print(f"{label:<20} {count:>6} {pct:>5.1f}%")

    # Sample prices by tier
    print(f"\n{'Cost':>8} {'Markup':>6} {'Raw':>8} {'Retail':>8} {'Margin':>7}")
    print("-" * 42)
    samples = [0.35, 1.50, 5.00, 10.00, 20.00, 45.00, 80.00, 200.00, 500.00]
    for cost in samples:
        markup = get_markup(cost)
        raw = cost * markup
        retail = calculate_retail_price(cost)
        margin = (retail - cost) / retail * 100
        print(f"${cost:>7.2f} {markup:>5.1f}x ${raw:>7.2f} ${retail:>7.2f} {margin:>5.1f}%")


if __name__ == "__main__":
    main()
