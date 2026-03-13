#!/usr/bin/env python3
"""
clean_names.py — Generate clean, customer-friendly product names for Shopify.

Steel City's raw product descriptions follow patterns like:
  "FILTER,HEPA-CLARKE COMFORT BACKPACK"
  "CORD-17-2 x 33' GRAY-BERNINA 6400/6600"
  "BRUSH ROLL-SIMPLICITY/RICCAR,F3700,R17,RSL3"

This script transforms them into clean e-commerce titles like:
  "Clarke HEPA Filter - Comfort Backpack"
  "Evolution 33' Gray Power Cord - Bernina 6400/6600"
  "Simplicity Brush Roll - F3700, R17, RSL3"
"""

import json
import os
import re

import openpyxl

SCHEMATIC_EXCEL = "output/steel_city_parts.xlsx"
CATALOG_EXCEL = "output/catalog_new_products.xlsx"
IMAGES_DIR = "images"
OUTPUT_JSON = "product_names.json"


# ─── Abbreviation Expansions ──────────────────────────────────────────────────

ABBREVIATIONS = {
    'ASSY': 'Assembly', 'ASSEM': 'Assembly', 'ASSEMBLE': 'Assembly',
    'ASY': 'Assembly', 'MTR': 'Motor', 'BRG': 'Bearing',
    'BLK': 'Black', 'WHT': 'White', 'GRN': 'Green', 'BLU': 'Blue',
    'GRY': 'Gray', 'SLVR': 'Silver', 'CLR': 'Clear', 'ORG': 'Orange',
    'PNK': 'Pink', 'PUR': 'Purple', 'YEL': 'Yellow', 'BRN': 'Brown',
    'REFL': 'Reflector', 'HNDL': 'Handle', 'VAC': 'Vacuum',
    'UPRT': 'Upright', 'UPRI': 'Upright', 'CLNR': 'Cleaner',
    'DIAM': 'Diameter', 'SQ': 'Square', 'PR': 'Pair',
    'COMMERICIAL': 'Commercial', 'CLOTHBAG': 'Cloth Bag',
}

KEEP_UPPER = {
    'HEPA', 'LED', 'UV', 'AC', 'DC', 'USA', 'XL', 'II', 'III', 'IV',
    'V', 'VI', 'VII', 'VIII', 'IX', 'RH', 'LH', 'PWR',
}


# ─── Core Cleaning ─────────────────────────────────────────────────────────────

def extract_raw_name(description):
    if not description:
        return ""
    desc = str(description).strip()
    if " - " in desc:
        return desc.split(" - ", 1)[1].strip()
    return desc


def is_model_number(word):
    """Check if a word looks like a model number (mix of letters and digits)."""
    clean = word.strip(',-/()')
    if not clean:
        return False
    has_digit = bool(re.search(r'\d', clean))
    has_alpha = bool(re.search(r'[A-Za-z]', clean))
    # Model numbers: UH70120, F5857, S3681D, BH50010, 86T3, etc.
    if has_digit and has_alpha and len(clean) >= 3:
        return True
    # Pure long numbers that look like model numbers: 1890, 2252
    if clean.isdigit() and len(clean) >= 4:
        return True
    return False


def clean_product_name(raw_name, brand, part_number):
    """Transform raw Steel City name into customer-friendly Shopify title."""
    name = raw_name
    if not name:
        return f"{brand} Replacement Part" if brand else "Replacement Part"

    first_brand = brand.split(",")[0].strip() if brand else ""
    all_brands = [b.strip() for b in brand.split(",") if b.strip()] if brand else []

    # ── Remove HTML ──
    name = re.sub(r'<br\s*/?>', ' | ', name, flags=re.IGNORECASE)
    name = re.sub(r'<[^>]+>', '', name)

    # ── Remove junk ──
    name = re.sub(r'\bNLA\b[\s\d/ -]*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'PLEASE USE ALT.*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'DO NOT USE.*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'USE ALT\s*#?\s*.*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\bDISCONTINUED\b[^|]*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\bOEM\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(REPL|REPLACES?)\s+[\w-]+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'SAME AS\s+[\w-]+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'SPARE PARTS', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\d+[A-Z]?\s*:\s*', '', name)

    # ── Remove brand names from body (we'll add as prefix) ──
    for b in all_brands:
        # Remove brand when it follows a separator or is standalone
        bp = re.escape(b)
        name = re.sub(rf'[-/,]\s*{bp}\b', ',', name, flags=re.IGNORECASE)
        name = re.sub(rf'^\s*{bp}\b\s*[,]?\s*', '', name, flags=re.IGNORECASE)
        # Also remove from middle: "FILTER,DIRT DEVIL F52" → "FILTER, F52"
        name = re.sub(rf'\b{bp}\b\s*,?\s*', '', name, flags=re.IGNORECASE)

    # ── Expand W/ → "with" ──
    name = re.sub(r'\bW\s*/\s*', 'with ', name, flags=re.IGNORECASE)
    name = re.sub(r'\bW/(?=\w)', 'with ', name, flags=re.IGNORECASE)

    # ── Expand N-Pack patterns ──
    name = re.sub(r'\((\d+)\s*PK\)', r'(\1 Pack)', name, flags=re.IGNORECASE)
    name = re.sub(r'\((\d+)\s*PAK\)', r'(\1 Pack)', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(\d+)\s*PK\b', r'\1 Pack', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(\d+)\s*PAK\b', r'\1 Pack', name, flags=re.IGNORECASE)
    name = re.sub(r'PACKS? OF (\d+)', r'\1 Pack', name, flags=re.IGNORECASE)

    # ── Expand abbreviations ──
    for abbr, expansion in ABBREVIATIONS.items():
        name = re.sub(rf'\b{abbr}\b', expansion, name, flags=re.IGNORECASE)

    # ── "ALSO FITS" → "Fits" ──
    name = re.sub(r'\bALSO FITS?\b', 'Fits', name, flags=re.IGNORECASE)

    # ── Restructure the comma/hyphen separated format ──
    # Split into segments on | (was <br>) and process each
    segments = [s.strip() for s in name.split('|') if s.strip()]

    # Process main segment (first one has the product name)
    main = segments[0] if segments else name
    extra = segments[1:] if len(segments) > 1 else []

    # Parse the main segment: typically "TYPE,MODIFIER1,MODIFIER2-CONTEXT"
    # or "TYPE-CONTEXT MODEL"
    main = restructure_name(main)

    # Rejoin with extra info
    if extra:
        extra_clean = [restructure_name(e) for e in extra]
        # Filter out empty extras
        extra_clean = [e for e in extra_clean if e and len(e) > 2]
        if extra_clean:
            main = main + " - " + ", ".join(extra_clean)

    name = main

    # ── Title case ──
    name = smart_title_case(name)

    # ── Add brand prefix ──
    if first_brand:
        name_lower = name.lower()
        brand_lower = first_brand.lower()
        if not name_lower.startswith(brand_lower):
            name = f"{first_brand} {name}"

    # ── Final cleanup ──
    name = re.sub(r'\s{2,}', ' ', name)
    name = re.sub(r',\s*,', ',', name)
    name = re.sub(r'\(\s*\)', '', name)
    name = re.sub(r'^\s*[-,/]\s*', '', name)
    name = re.sub(r'\s*[-,/]\s*$', '', name)
    name = re.sub(r'\s*,\s*$', '', name)
    name = re.sub(r'\s+-\s*$', '', name)
    name = name.strip()

    # Cap at 120 chars
    if len(name) > 120:
        # Try to cut at comma or dash
        for sep in [' - ', ', ']:
            cut = name[:120].rfind(sep)
            if cut > 50:
                name = name[:cut]
                break
        else:
            cut = name[:120].rfind(' ')
            if cut > 50:
                name = name[:cut]
        name = name.strip(' ,-/')

    return name


def restructure_name(text):
    """
    Restructure Steel City's comma-separated format into readable English.

    "FILTER,HEPA,EXHAUST" → "HEPA Exhaust Filter"
    "CORD,50' 16-3 ORANGE" → "50' Orange Power Cord"
    "BRUSH ROLL" → "Brush Roll" (no change needed)
    """
    text = text.strip(' ,-/')
    if not text:
        return ""

    # Don't restructure if it's already fairly clean (no commas, or just model listing)
    if ',' not in text and '-' not in text:
        return text

    # Split on commas
    parts = [p.strip() for p in text.split(',') if p.strip()]

    if len(parts) <= 1:
        # No commas — just clean up hyphens used as separators
        # "FILTER-MODEL" → "Filter - Model" (keep hyphen as dash)
        return text

    # First part is usually the product type: FILTER, CORD, BRUSH ROLL, etc.
    product_type = parts[0].strip()

    # Remaining parts are modifiers/context
    modifiers = parts[1:]

    # Rebuild: put modifiers before type for more natural English
    # "FILTER, HEPA, EXHAUST" → "HEPA Exhaust Filter"
    # "CORD, 50' 16-3 ORANGE" → "50' 16-3 Orange Cord"
    # But "BRUSH ROLL, 13 1/8 INCHES" → "Brush Roll, 13 1/8 Inches"

    # Heuristic: if modifiers are short adjectives/descriptors, put before type
    # If modifiers are model numbers or long phrases, put after with dash
    short_mods = []
    long_mods = []

    for mod in modifiers:
        mod = mod.strip()
        # "Fits ..." always goes at end
        if mod.lower().startswith('fits'):
            long_mods.append(mod)
        # Model numbers go after
        elif is_model_number(mod.split()[0] if mod.split() else mod):
            long_mods.append(mod)
        # Short single-word descriptors go before (colors, types)
        elif len(mod.split()) <= 2 and not any(c.isdigit() for c in mod):
            short_mods.append(mod)
        else:
            long_mods.append(mod)

    # Build result
    result_parts = []
    if short_mods:
        result_parts.extend(short_mods)
    result_parts.append(product_type)

    result = ' '.join(result_parts)

    if long_mods:
        result += ' - ' + ', '.join(long_mods)

    return result


def smart_title_case(text):
    """Title case preserving model numbers and abbreviations."""
    words = text.split()
    result = []
    small_words = {'a', 'an', 'the', 'and', 'or', 'for', 'of', 'with', 'in', 'on', 'to', 'at', 'by'}

    for i, word in enumerate(words):
        stripped = word.strip(',-/()"\'')

        if not stripped:
            result.append(word)
            continue

        # Keep model numbers as-is (but uppercase them)
        if is_model_number(stripped):
            result.append(word.upper())
            continue

        # Keep known abbreviations uppercase
        if stripped.upper() in KEEP_UPPER:
            result.append(word.upper())
            continue

        # Pure numbers — keep as-is
        if re.match(r'^[\d\'"/.]+$', stripped):
            result.append(word)
            continue

        # Small words (not first word, not after a dash)
        if stripped.lower() in small_words and i > 0:
            prev = words[i - 1] if i > 0 else ''
            if not prev.endswith('-'):
                result.append(word.lower())
                continue

        # Standard title case
        result.append(word.capitalize())

    return ' '.join(result)


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    products = {}

    print("Loading schematics Excel...")
    wb = openpyxl.load_workbook(SCHEMATIC_EXCEL, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        pn = str(d.get("Part Number", "")).strip()
        if pn in products:
            continue
        if not os.path.exists(f"{IMAGES_DIR}/{pn}/1.jpg"):
            continue

        desc = str(d.get("Description", "") or "").strip()
        brand = str(d.get("Brand", "") or "").strip()
        sku = str(d.get("SKU", "") or "").strip()
        price = str(d.get("Price", "") or "").strip()
        model = str(d.get("Model", "") or "").strip()

        raw_name = extract_raw_name(desc)
        clean_name = clean_product_name(raw_name, brand, pn)

        products[pn] = {
            "raw_description": desc,
            "raw_name": raw_name,
            "clean_name": clean_name,
            "brand": brand,
            "sku": sku,
            "price": price,
            "model": model,
            "source": "schematics",
        }
    wb.close()
    print(f"  Schematics: {len(products)} parts with images")

    print("Loading catalog Excel...")
    wb2 = openpyxl.load_workbook(CATALOG_EXCEL, read_only=True)
    ws2 = wb2.active
    headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]

    catalog_count = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers2, row))
        pn = str(d.get("Part Number", "")).strip()
        if pn in products:
            continue
        if not os.path.exists(f"{IMAGES_DIR}/{pn}/1.jpg"):
            continue

        desc = str(d.get("Description", "") or "").strip()
        brand = str(d.get("Manufacturer", "") or "").strip()
        sku = str(d.get("SKU", "") or "").strip()
        price = str(d.get("Price", "") or "").strip()

        raw_name = extract_raw_name(desc) if " - " in desc else desc
        if not raw_name:
            raw_name = desc

        clean_name = clean_product_name(raw_name, brand, pn)

        products[pn] = {
            "raw_description": desc,
            "raw_name": raw_name,
            "clean_name": clean_name,
            "brand": brand,
            "sku": sku,
            "price": price,
            "model": "",
            "source": "catalog",
        }
        catalog_count += 1
    wb2.close()
    print(f"  Catalog: {catalog_count} additional parts with images")

    with open(OUTPUT_JSON, "w") as f:
        json.dump(products, f, indent=2)

    print(f"\nTotal: {len(products)} products with clean names → {OUTPUT_JSON}")

    # Print samples
    print("\n" + "=" * 80)
    print("SAMPLE NAMES (50 random)")
    print("=" * 80)
    import random
    random.seed(42)
    sample_keys = random.sample(list(products.keys()), min(50, len(products)))
    for pn in sorted(sample_keys):
        info = products[pn]
        print(f"  Raw:   {info['raw_name'][:90]}")
        print(f"  Clean: {info['clean_name']}")
        print()


if __name__ == "__main__":
    main()
